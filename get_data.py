#注意文件名为：get_data

from openpyxl import load_workbook


def get_excel_date():
#读取excle文件名
    workbook = load_workbook("一季度.xlsx")
#读取excel内的sheet
    worksheet = workbook["一季度汇总表"]

    data = []
#读取每一行？？？【待理解】
    for row in worksheet.iter_rows(min_row=1):
#创建一个名称为dict_data的空白字典
        dict_data = {}
#每一行第一个的数据赋值给name
        dict_data["name"] = row[0].value
#每一行的第三个数据复制给value
#python从0开始算
        dict_data["value"] = row[2].value
#将dict_data保存到data
        data.append(dict_data)
#关闭excel表格
    workbook.close()
#返回data数值
    return data
# note 输出数据
# print(get_excel_date())
